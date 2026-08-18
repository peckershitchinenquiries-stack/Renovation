#!/usr/bin/env python3
"""
RenovaTrack — prove the web app will show what the spreadsheet shows.

This reads the rows out of the GENERATED migration (not out of the spreadsheet
again), replays the app's own arithmetic on them exactly as lib/calculations.ts
and lib/summary.ts do, and compares the result against
46_Glenferrie_Rd_Renovation_Spend_Tracker_Updated.xlsx.

Verifying the SQL rather than the parser is the point: it checks the artifact
that actually gets pasted into Supabase.

What it checks
  1. every diary row's total incl. VAT vs the sheet's 'Total incl. VAT' cell
  2. every week's total vs the Summary sheet's 'By Week (Totals)' block
  3. the Labour / Materials split per week
  4. the seven Overview cards vs the Summary sheet's Controls and
     Payment Status blocks — Target Budget included, which must equal the
     sheet's own (blank) Target Budget cell
  5. row counts and paid-row counts, and that the ledger count is 0:
     Renovation_Cost_Tracker-1.xlsx is a different job and must not reappear

Usage:
    python scripts/verify_against_spreadsheet.py
Exit code 0 = everything matches, 1 = at least one check failed.
"""
import re
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
SQL = ROOT / "supabase" / "migrations" / "0009_reimport_file1_only.sql"
XLSX = ROOT / "46_Glenferrie_Rd_Renovation_Spend_Tracker_Updated.xlsx"

TOL = 0.01          # a penny: anything smaller is float/rounding noise
failures = []


def check(name, got, want, tol=TOL, fmt="{:,.2f}"):
    ok = abs(got - want) <= tol
    if not ok:
        failures.append(f"{name}: app {fmt.format(got)} vs sheet {fmt.format(want)}")
    mark = "PASS" if ok else "FAIL"
    print(f"  [{mark}] {name:<34} app {fmt.format(got):>14}   sheet {fmt.format(want):>14}")
    return ok


# ------------------------------------------------------------ read the SQL
def split_top_level(s):
    """Split a SQL tuple body on commas that are not inside a quoted string."""
    out, buf, in_str, i = [], [], False, 0
    while i < len(s):
        c = s[i]
        if in_str:
            if c == "'":
                if i + 1 < len(s) and s[i + 1] == "'":   # '' escape
                    buf.append("''")
                    i += 2
                    continue
                in_str = False
            buf.append(c)
        elif c == "'":
            in_str = True
            buf.append(c)
        elif c == ",":
            out.append("".join(buf).strip())
            buf = []
        else:
            buf.append(c)
        i += 1
    out.append("".join(buf).strip())
    return out


def unquote(tok):
    tok = tok.strip()
    if tok == "null":
        return None
    tok = re.sub(r"::\w+$", "", tok)
    if tok.startswith("'") and tok.endswith("'"):
        return tok[1:-1].replace("''", "'")
    return tok


FIELDS = [
    "user", "project", "week", "description", "category", "trade", "room",
    "notes", "supplier", "invoice", "paid_date", "method", "qty", "unit_cost",
    "vat", "status", "quoted", "actual", "paid", "source",
]


def load_sql_rows():
    """Pull the expense_entries VALUES tuples out of the generated migration."""
    text = SQL.read_text(encoding="utf-8")
    start = text.index("insert into public.expense_entries")
    body = text[start:text.index(";", start)]
    rows = []
    for line in body.splitlines():
        line = line.strip().rstrip(",")
        if not (line.startswith("(v_user,") and line.endswith(")")):
            continue
        parts = split_top_level(line[1:-1])
        if len(parts) != len(FIELDS):
            raise SystemExit(f"unexpected column count {len(parts)} in: {line[:80]}")
        r = dict(zip(FIELDS, (unquote(p) for p in parts)))
        for k in ("week", "qty", "unit_cost", "vat", "quoted", "actual", "paid"):
            r[k] = float(r[k])
        r["week"] = int(r["week"])
        rows.append(r)
    return rows


def sql_target_budget():
    """The target_budget the migration will write into public.projects.

    Read out of the generated SQL rather than recomputed, so this checks the
    artifact that actually gets pasted into Supabase.
    """
    text = SQL.read_text(encoding="utf-8")
    m = re.search(r"values \(v_user, '[^']*', ([\d.]+), 'active'", text)
    if not m:
        raise SystemExit("could not find the projects insert in " + SQL.name)
    return float(m.group(1))


# ------------------------- the app's arithmetic, mirrored from the TS source
def compute(r):
    """lib/calculations.ts -> computeEntry()."""
    vat_amount = r["actual"] * (r["vat"] / 100)
    total = r["actual"] + vat_amount
    return {**r, "vat_amount": vat_amount, "total_incl_vat": total,
            "remaining": total - r["paid"]}


def build_summary(entries, target_budget):
    """lib/summary.ts -> buildSummary()."""
    active = [e for e in entries if e["status"] != "Cancelled"]
    total_quoted = sum(e["quoted"] for e in active)
    forecast = sum(e["total_incl_vat"] for e in active)
    paid = sum(e["paid"] for e in active)
    variance = round((forecast - total_quoted) * 100) / 100 or 0.0
    return {
        "target_budget": target_budget,
        "total_quoted": total_quoted,
        "forecast_total": forecast,
        "variance": variance,
        "paid_to_date": paid,
        "remaining_to_pay": forecast - paid,
        "weeks_tracked": len({e["week"] for e in active}),
    }


def build_by_week(entries):
    """lib/summary.ts -> buildByWeek()."""
    out = {}
    for e in entries:
        if e["status"] == "Cancelled":
            continue
        w = out.setdefault(e["week"], {"labour": 0.0, "materials": 0.0,
                                       "vat": 0.0, "total": 0.0, "rows": 0})
        if e["category"] == "Materials":
            w["materials"] += e["total_incl_vat"]
        else:
            w["labour"] += e["total_incl_vat"]
        w["vat"] += e["vat_amount"]
        w["total"] += e["total_incl_vat"]
        w["rows"] += 1
    return out


# ------------------------------------------------------------ read the sheet
def load_sheet():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    plan = wb["Week-by-Week Plan"]
    rows = []
    for r in list(plan.iter_rows(values_only=True))[5:]:
        try:
            week = int(str(r[0]).strip())
        except (TypeError, ValueError):
            continue
        desc = str(r[1]).strip() if r[1] else ""
        if not desc:
            continue
        rows.append({"week": week, "description": desc,
                     "total": round(float(r[17] or 0), 2),
                     "category": str(r[2]).strip() if r[2] else None,
                     "paid_date": r[8]})

    summ = wb["Summary"]
    by_week, controls = {}, {}
    for r in summ.iter_rows(values_only=True):
        if isinstance(r[0], (int, float)) and r[3] is not None:
            by_week[int(r[0])] = {
                "labour": float(r[1] or 0),
                "materials": float(r[2] or 0),
                "total": float(r[3] or 0),
            }
        elif isinstance(r[0], str) and isinstance(r[1], (int, float)):
            controls[r[0].strip()] = float(r[1])
    return rows, by_week, controls


# ------------------------------------------------------------------- main
def main():
    sql_rows = load_sql_rows()
    diary = [compute(r) for r in sql_rows if r["source"] == "diary"]
    ledger = [r for r in sql_rows if r["source"] == "ledger"]
    sheet_rows, sheet_weeks, controls = load_sheet()

    print(f"source SQL   : {SQL.name}")
    print(f"spreadsheet  : {XLSX.name}")
    print(f"diary rows   : {len(diary)}   ledger rows: {len(ledger)}\n")

    # -- 1. counts ---------------------------------------------------------
    print("1. Row counts")
    check("diary rows", len(diary), len(sheet_rows), tol=0, fmt="{:,.0f}")
    # Renovation_Cost_Tracker-1.xlsx is a different job and is no longer
    # imported; anything here means it has crept back in.
    check("ledger rows", len(ledger), 0, tol=0, fmt="{:,.0f}")
    sheet_paid = sum(1 for r in sheet_rows if r["paid_date"])
    app_paid = sum(1 for e in diary if e["paid"] > 0)
    check("rows marked Paid", app_paid, sheet_paid, tol=0, fmt="{:,.0f}")

    # -- 2. every row's total ---------------------------------------------
    print("\n2. Per-row total incl. VAT (all rows, in sheet order)")
    bad = 0
    for e, s in zip(diary, sheet_rows):
        if e["week"] != s["week"] or e["description"] != s["description"]:
            failures.append(
                f"row order drift: SQL has week {e['week']} {e['description']!r}, "
                f"sheet has week {s['week']} {s['description']!r}")
            bad += 1
            continue
        if abs(e["total_incl_vat"] - s["total"]) > TOL:
            failures.append(
                f"week {e['week']} {e['description']!r}: app "
                f"{e['total_incl_vat']:.2f} vs sheet {s['total']:.2f}")
            bad += 1
    print(f"  [{'PASS' if bad == 0 else 'FAIL'}] {len(diary) - bad} of {len(diary)} rows match "
          f"to the penny")

    # -- 3. weekly totals and the labour/materials split -------------------
    print("\n3. Week-by-Week table (chart + Overview table)")
    by_week = build_by_week(diary)

    # The app's Labour/Materials split is 'Category = Materials vs everything
    # else', on incl-VAT totals. Derive the same thing straight from the sheet's
    # own rows so the split is checked against the source, not against the
    # Summary tab's columns (which measure something different — see the note).
    want_split = {}
    for s in sheet_rows:
        b = want_split.setdefault(s["week"], {"labour": 0.0, "materials": 0.0})
        key = "materials" if s["category"] == "Materials" else "labour"
        b[key] += s["total"]

    print(f"  {'wk':>3} {'rows':>5} {'app total':>13} {'sheet total':>13}"
          f" {'app labour':>12} {'app mat':>12}   split")
    bad_total = bad_split = 0
    for w in sorted(by_week):
        s = sheet_weeks.get(w, {"labour": 0, "materials": 0, "total": 0})
        a, want = by_week[w], want_split.get(w, {"labour": 0.0, "materials": 0.0})
        ok = abs(a["total"] - s["total"]) <= TOL
        split_ok = (abs(a["labour"] - want["labour"]) <= TOL
                    and abs(a["materials"] - want["materials"]) <= TOL)
        if not ok:
            bad_total += 1
            failures.append(f"week {w} total: app {a['total']:.2f} vs sheet {s['total']:.2f}")
        if not split_ok:
            bad_split += 1
            failures.append(f"week {w} split: app labour {a['labour']:.2f} / "
                            f"materials {a['materials']:.2f} vs sheet rows "
                            f"{want['labour']:.2f} / {want['materials']:.2f}")
        print(f"  {w:>3} {a['rows']:>5} {a['total']:>13,.2f} {s['total']:>13,.2f}"
              f" {a['labour']:>12,.2f} {a['materials']:>12,.2f}"
              f"   {'ok' if split_ok else 'FAIL'}{'' if ok else '   <-- FAIL'}")
    print(f"  [{'PASS' if bad_total == 0 else 'FAIL'}] all {len(by_week)} week totals match")
    print(f"  [{'PASS' if bad_split == 0 else 'FAIL'}] all {len(by_week)} Labour/Materials "
          f"splits match the sheet's rows")

    # Informational: the Summary tab's own two columns will not agree with the
    # app, for two reasons that are both properties of the spreadsheet.
    print("\n  For reference, the Summary tab's own Labour/Materials columns:")
    print(f"  {'wk':>3} {'sheet lab':>12} {'sheet mat':>12}   differs because")
    for w in sorted(by_week):
        s = sheet_weeks.get(w, {"labour": 0, "materials": 0})
        a = by_week[w]
        why = ""
        if abs(a["labour"] - s["labour"]) > TOL or abs(a["materials"] - s["materials"]) > TOL:
            # Summary reads the ex-VAT cost columns; the app shows incl-VAT.
            # On weeks 20-23 every cost was also typed into the Materials
            # column regardless of the row's Category.
            why = "cost typed in the Materials column" if s["labour"] == 0 and a["labour"] > 0 \
                else "Summary columns are ex-VAT, the app shows incl-VAT"
        print(f"  {w:>3} {s['labour']:>12,.2f} {s['materials']:>12,.2f}   {why}")
    print("  Neither affects a week total, a card, or the chart's bar heights.")

    # -- 4. the Overview cards --------------------------------------------
    print("\n4. Overview cards vs the Summary sheet")
    summ = build_summary(diary, sql_target_budget())
    forecast_sheet = controls.get("Forecast Total (incl. VAT)", 0.0)
    paid_sheet = controls.get("Paid to date (£)", 0.0)
    committed_sheet = controls.get("Committed / no paid-date logged (£)", 0.0)

    check("Actual Total (incl. VAT)", summ["forecast_total"], forecast_sheet)
    check("Total Quoted", summ["total_quoted"], forecast_sheet)
    check("Variance vs Quote", summ["variance"], 0.0)
    check("Paid to Date", summ["paid_to_date"], paid_sheet)
    check("Remaining to Pay", summ["remaining_to_pay"], committed_sheet)
    # The Summary tab lists weeks 1-32; the ones past 23 are still empty and the
    # app only counts weeks that actually carry an entry.
    weeks_with_spend = sum(1 for v in sheet_weeks.values() if v["total"] > 0)
    check("Weeks Tracked", summ["weeks_tracked"], weeks_with_spend, tol=0, fmt="{:,.0f}")
    budget_sheet = controls.get("Target Budget (incl. VAT)", 0.0)
    check("Target Budget", summ["target_budget"], budget_sheet)
    if summ["target_budget"] == 0:
        print("         (blank in the sheet and 0 here — the card and the "
              "'% of budget' bar hide themselves)")

    # -- 5. category donut -------------------------------------------------
    print("\n5. Category donut")
    mats = sum(e["total_incl_vat"] for e in diary
               if e["category"] == "Materials" and e["status"] != "Cancelled")
    lab = sum(e["total_incl_vat"] for e in diary
              if e["category"] != "Materials" and e["status"] != "Cancelled")
    sheet_by_cat_mats = sum(
        s["total"] for s in sheet_rows if s["category"] == "Materials")
    sheet_by_cat_lab = sum(
        s["total"] for s in sheet_rows if s["category"] != "Materials")
    check("Materials", mats, sheet_by_cat_mats)
    check("Labour (everything else)", lab, sheet_by_cat_lab)

    # -- verdict -----------------------------------------------------------
    print("\n" + "=" * 74)
    if failures:
        print(f"FAILED — {len(failures)} check(s) did not match:")
        for f in failures:
            print("  - " + f)
        return 1
    print("ALL CHECKS PASSED — the web app will show exactly what the "
          "spreadsheet shows.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
