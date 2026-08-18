#!/usr/bin/env python3
"""
RenovaTrack — rebuild the Supabase data from the source spreadsheet.

Generates supabase/migrations/0009_reimport_file1_only.sql, which recreates the
project, its expense_entries and its trade_lookups under the CURRENT auth user.

ONE source file, and that is deliberate:
  46_Glenferrie_Rd_Renovation_Spend_Tracker_Updated.xlsx
    'Week-by-Week Plan' -> 111 diary rows, weeks 1-23
    'Lookups'           -> 16 trade lookups

WHY Renovation_Cost_Tracker-1.xlsx IS NO LONGER IMPORTED  (2026-08-14)
----------------------------------------------------------------------
Up to 0007 that second workbook was imported as 96 'ledger' rows. It is a
different job, not a second record of this one:

  - it names no address; File 1 names 46 Glenferrie Road on both sheets
  - its dates run 2025-11-02 to 2026-01-25 and STOP a month before File 1's
    week 1 (2026-02-27). The two ranges do not overlap by a single day
  - its last entries are carpets, staging and a driveway clean — a finished
    house — while File 1's week 1 is 'clearance / back to brick'
  - every one of its rows is 100% paid; File 1 has paid GBP 13,273 of 151,645

It also poisoned the Price Tracker. Its 'Materials & Suppliers' sheet is a
payment log: the Item column is empty on all 60 rows, Quantity is 1 on all 60,
and Unit Cost holds THE AMOUNT OF THAT PAYMENT. So a deposit of 416.05 followed
by a balance of 3,586.00 to Wunda UFH read as one item whose unit price rose
761.9%. Eleven suppliers were doing this. Since File 1 records no unit costs at
all, that meant 100% of the Price Tracker's content was instalment payments
misread as prices.

The workbook stays in the repo as history. It is not deleted, just not imported.

HOW MONEY IS MAPPED  (read this before changing anything)
--------------------------------------------------------
The app never stores a total. `computeEntry` in lib/calculations.ts derives

    total_incl_vat = actual_amount x (1 + vat_rate / 100)

on every read. So `actual_amount` MUST hold the EX-VAT figure, or VAT gets
applied twice. The earlier import stored the sheet's 'Total incl. VAT (£)'
column in actual_amount *and* set vat_rate to 20, which is exactly what made
the app show £43,686.17 where the spreadsheet said £42,411.81.

    actual_amount  <- 'Labour Cost (£)' + 'Materials Cost (£)'   (ex-VAT)
    vat_rate       <- 'VAT'  ('0%' -> 0, '20%' -> 20)
    quoted_amount  <- 'Total incl. VAT (£)'
    paid_amount    <- 'Total incl. VAT (£)' when a Paid Date is filled, else 0
    status         <- 'Paid' when a Paid Date is filled, else the sheet's Status

`quoted_amount` gets the incl-VAT total because the Week-by-Week Plan has no
quote column at all — the sheet's forecast IS its quote. That makes the
Overview 'Total Quoted' card agree with the sheet's 'Forecast Total (incl.
VAT)' and leaves 'Variance vs Quote' at £0.

`paid_amount` gets the incl-VAT total because that is what was actually handed
over, and it makes 'Paid to Date' and 'Remaining to Pay' agree with the sheet's
own Payment Status block to the penny.

Every row is checked against the sheet's own arithmetic before it is emitted
(see check_arithmetic), and the generated SQL re-checks every week total
against the Summary sheet before it commits.

Usage:
    python scripts/build_import_sql.py
"""
import re
import datetime as dt
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
USER_ID = "5d3fc9ff-92a3-4923-a18b-7eb5eade3105"  # admin@pk.com, current account
PROJECT_NAME = "46 Glenferrie Road"

FILE1 = ROOT / "46_Glenferrie_Rd_Renovation_Spend_Tracker_Updated.xlsx"
OUT = ROOT / "supabase" / "migrations" / "0009_reimport_file1_only.sql"

# The spreadsheet's own 'Target Budget (incl. VAT)' cell is blank, so there is
# no budget to import. 0 is the honest value: OverviewTab, ProjectDetail,
# the dashboard and the project list all hide the budget card and the
# '% of budget' bar when it is 0, rather than showing a ratio against a figure
# that came from somewhere else. Set it in the app's Edit Project form if a
# real ceiling is ever agreed.
TARGET_BUDGET = 0.0

# Column indexes into 'Week-by-Week Plan' (0-based, header on sheet row 5).
C_WEEK, C_DESC, C_CAT, C_TRADE, C_ROOM = 0, 1, 2, 3, 4
C_NOTES, C_SUPPLIER, C_INVOICE, C_PAIDDATE, C_METHOD = 5, 6, 7, 8, 9
C_HOURS, C_RATE, C_LABOUR = 10, 11, 12
C_QTY, C_UNIT, C_MATERIALS = 13, 14, 15
C_VAT, C_TOTAL, C_STATUS = 16, 17, 18

# Values permitted by the CHECK constraints in 0001_init.sql.
CATEGORIES = {"Labour", "Materials", "Skip/Disposal", "Other"}
STATUSES = {"Planned", "In Progress", "Paid", "Cancelled"}
METHODS = {"Cash", "Debit Card", "Credit Card", "Bank Transfer"}

PENNY = 0.011  # tolerance for float comparisons against the sheet


# --------------------------------------------------------------- suppliers
# WHERE THE MERCHANT NAMES ACTUALLY LIVE  (2026-08-17)
# ----------------------------------------------------------
# The sheet's 'Supplier' column is not usable. It has only 8 non-empty cells
# and 7 of them are sentences typed into the wrong column ('steels in',
# '1 day - to DPC', '£300 PAID FROM OWED'). The merchant names were being
# typed into 'Task / Description' instead — so before this change /suppliers
# listed seven notes and one real merchant, and every real merchant was
# invisible.
#
# The list below was confirmed by the project owner. It is exactly the set of
# distinct descriptions on the sheet's 45 Materials rows — check_suppliers()
# proves that on every run and ABORTS if the two ever drift apart. So this is
# a declared mapping, not a heuristic: nothing here guesses which strings
# "look like" a merchant, and a new Materials row with an unlisted description
# stops the import rather than silently inventing a supplier.
#
# Kept verbatim as the owner wrote them, including the two pairs that are
# probably one merchant each — 'Johnstones' / 'Johnstones Paint' and
# 'Steels' / 'Ryan Steels'. Merging near-duplicates is the alias work
# described in about.md §4.6, where a human confirms each one; it is not this
# script's job to decide.
SUPPLIER_NAMES = [
    "Master Mix",
    "Steels",
    "SDE Drainage",
    "Scaffold",
    "Lawsons",
    "Eurocell",
    "Wunda UFH Materials",
    "Cabinets Direct (Kitchen)",
    "Alspec Windows",
    "Miscl Roofing Materials",
    "Travis Perkins",
    "CAD Stairs",
    "St Albans Bathroom Centre",
    "Saris",
    "Todds Doors",
    "Ryan Steels",
    "Eaves Electrical",
    "Johnstones Paint",
    "JJ Roofing",
    "Jewsons",
    "Watts Roofing",
    "Pro Tiler",
    "Skips",
    "Cut Price Tiles",
    "Plastic Construction",
    "Metres Direct LTD",
    "Corston",
    "Lionvest",
    "Topps Tiles",
    "AC Supplies",
    "Johnstones",
    "Mark Cornice",
]

# The one Supplier-column cell that really is a merchant: sheet row 59,
# week 5, description 'Skip'. It is a Labour row, so the Materials rule below
# would never reach it. Every other non-empty cell in that column is a note.
SUPPLIER_COLUMN_MERCHANTS = ["Stevenage Skips"]


# ----------------------------------------------------------------- helpers
def q(v):
    """Quote a value as a SQL literal."""
    if v is None or v == "":
        return "null"
    return "'" + str(v).replace("'", "''") + "'"


def num(v, default=0):
    """Coerce a spreadsheet cell to a number, tolerating '£1,234.50' text."""
    if v is None or v == "":
        return default
    if isinstance(v, (int, float)):
        return round(float(v), 2)
    s = re.sub(r"[^0-9.\-]", "", str(v))
    try:
        return round(float(s), 2)
    except ValueError:
        return default


def text(v):
    """Trimmed cell text, or None when the cell is empty."""
    s = str(v).strip() if v is not None else ""
    return s or None


WEEKDAYS = {
    "monday": 0, "tuesday": 1, "wednesday": 2, "thursday": 3,
    "friday": 4, "saturday": 5, "sunday": 6,
}
# Candidate years for resolving free-text paid dates, most likely first.
YEAR_CANDIDATES = (2026, 2025, 2027)


def as_date(v):
    """Return an ISO date string, or None if the cell isn't a real date."""
    if isinstance(v, (dt.datetime, dt.date)):
        return v.strftime("%Y-%m-%d")
    return None


def resolve_written_date(v):
    """Resolve free text like 'Friday 27/2' to a real ISO date.

    File 1's 'Paid Date' column is hand-typed day/month with a weekday name and
    no year. The weekday disambiguates the year: only 2026 puts every one of
    these dates on the stated weekday. Returns None if it can't be resolved,
    in which case the raw text is kept in notes as before.
    """
    if isinstance(v, (dt.datetime, dt.date)):
        return None  # already a real date; as_date handles it
    s = str(v or "").strip()
    if not s:
        return None
    m = re.search(r"(\d{1,2})\s*/\s*(\d{1,2})", s)
    if not m:
        return None
    day, month = int(m.group(1)), int(m.group(2))
    named = next((n for n in WEEKDAYS if n in s.lower()), None)

    fallback = None
    for year in YEAR_CANDIDATES:
        try:
            cand = dt.date(year, month, day)
        except ValueError:
            continue
        if named is None:
            return cand.strftime("%Y-%m-%d")
        if cand.weekday() == WEEKDAYS[named]:
            return cand.strftime("%Y-%m-%d")
        fallback = fallback or cand
    # Weekday never matched — trust the day/month over the weekday name.
    return fallback.strftime("%Y-%m-%d") if fallback else None


def category_of(v):
    v = (str(v).strip() if v else "")
    return v if v in CATEGORIES else ("Other" if v else None)


def status_of(v, paid):
    """Map a spreadsheet status to the CHECK-constrained set."""
    s = str(v or "").strip()
    if "Cancel" in s:
        return "Cancelled"
    if "Progress" in s:
        return "In Progress"
    if paid > 0:
        return "Paid"
    if s in STATUSES:
        return s
    return "Planned"


def method_of(v):
    s = str(v or "").strip()
    if not s:
        return None
    low = s.lower()
    if "cash" in low:
        return "Cash"
    if "credit" in low:
        return "Credit Card"
    if "debit" in low:
        return "Debit Card"
    if "bank" in low or "transfer" in low or low.startswith("bt"):
        return "Bank Transfer"
    return None


def vat_of(v):
    """vat_rate is constrained to exactly 0 or 20.

    The updated sheet writes this column as text ('0%' / '20%'); the older one
    wrote it as a number. num() strips the '%' either way.
    """
    n = num(v)
    if n in (0, 20):
        return int(n)
    return 20 if n > 0 else 0


def name_key(v):
    """Trim, lower-case, collapse internal whitespace.

    Deliberately the same rule as public.norm_key() in 0008, priceKey() in
    lib/summary.ts and normaliseName() in lib/purchases.ts. All four must stay
    in step or the import, the database and the app will disagree about what
    one merchant is. It is what lets 'CAD Stairs ' (trailing space on the
    sheet) resolve to the canonical 'CAD Stairs'.
    """
    return re.sub(r"\s+", " ", str(v or "").strip().lower())


SUPPLIER_BY_KEY = {
    name_key(n): n for n in SUPPLIER_NAMES + SUPPLIER_COLUMN_MERCHANTS
}


def supplier_of(description, supplier_cell, category):
    """Resolve a row's merchant, and rescue anything misfiled in Supplier.

    Returns (supplier, misplaced_note):
      - supplier       the canonical merchant name, or None for a Labour row
                       that has no merchant behind it
      - misplaced_note the Supplier cell's text when it is not a merchant —
                       the caller folds it into notes, which is the column it
                       should have been typed into in the first place

    Order matters: a Supplier cell naming a known merchant wins, because that
    is the column actually meant for it. Otherwise a Materials row takes its
    merchant from the description, which is where they were all typed.
    """
    cell = text(supplier_cell)
    if cell and name_key(cell) in SUPPLIER_BY_KEY:
        return SUPPLIER_BY_KEY[name_key(cell)], None
    named = SUPPLIER_BY_KEY.get(name_key(description)) if category == "Materials" else None
    return named, cell


def join_notes(*parts):
    out = [str(p).strip() for p in parts if p not in (None, "")]
    return " | ".join(out) if out else None


COLS = (
    "user_id, project_id, week_number, description, category, trade, "
    "location_room, notes, supplier, invoice_ref, paid_date, payment_method, "
    "qty, unit_cost, vat_rate, status, quoted_amount, actual_amount, "
    "paid_amount, source"
)


def row_sql(r):
    return (
        "  ("
        f"v_user, v_project, {r['week']}, {q(r['description'])}, "
        f"{q(r['category'])}, {q(r['trade'])}, {q(r['location'])}, "
        f"{q(r['notes'])}, {q(r['supplier'])}, {q(r['invoice'])}, "
        f"{q(r['paid_date'])}{'::date' if r['paid_date'] else ''}, "
        f"{q(r['method'])}, {r['qty']}, {r['unit_cost']}, {r['vat']}, "
        f"{q(r['status'])}, {r['quoted']}, {r['actual']}, {r['paid']}, "
        f"{q(r['source'])})"
    )


# ----------------------------------------------------------------- File 1
def diary_sheet():
    return openpyxl.load_workbook(FILE1, data_only=True)["Week-by-Week Plan"]


def parse_diary():
    """Week-by-Week Plan -> diary rows (the in-app Expenses tab).

    Returns (rows, warnings). A row is real when the week cell is a number AND
    the description is non-empty; everything else is a 'Week n' banner or one
    of the blank padding rows the template ships with.
    """
    rows = list(diary_sheet().iter_rows(values_only=True))
    out, warnings = [], []

    for excel_row, r in enumerate(rows[5:], start=6):
        try:
            week = int(str(r[C_WEEK]).strip())
        except (TypeError, ValueError):
            continue  # 'Week 1' banner, or blank
        description = text(r[C_DESC])
        if not description:
            continue  # padding row: a week number with no task

        raw_date = r[C_PAIDDATE]
        # A filled 'Paid Date' is what marks a row as paid — the sheet's Status
        # column reads 'Planned' on every row and carries no payment signal.
        paid_date = as_date(raw_date) or resolve_written_date(raw_date)
        if raw_date and not paid_date:
            warnings.append(
                f"row {excel_row} (week {week}, {description!r}): could not read "
                f"paid date {raw_date!r} — kept as text in notes"
            )
        is_paid = bool(paid_date)

        # Ex-VAT cost. The sheet splits it across two columns; a Materials row
        # carries its money in 'Materials Cost', a Labour row in 'Labour Cost'.
        # Weeks 20-23 were typed with everything in the Materials column, which
        # is why we add the pair rather than choosing one by category.
        ex_vat = round(num(r[C_LABOUR]) + num(r[C_MATERIALS]), 2)
        vat = vat_of(r[C_VAT])
        sheet_total = num(r[C_TOTAL])

        category = category_of(r[C_CAT])
        # The merchant is in the description on Materials rows, not in the
        # Supplier column — see SUPPLIER_NAMES above. Anything else sitting in
        # the Supplier column is a note and is moved into notes.
        supplier, misplaced = supplier_of(description, r[C_SUPPLIER], category)
        if misplaced:
            warnings.append(
                f"row {excel_row} (week {week}, {description!r}): Supplier cell "
                f"{misplaced!r} is not a merchant — moved into notes"
            )

        out.append({
            "week": week,
            "excel_row": excel_row,
            "description": description,
            "category": category,
            "trade": text(r[C_TRADE]),
            "location": text(r[C_ROOM]),
            # Preserve the unparseable date text rather than losing it.
            "notes": join_notes(
                r[C_NOTES],
                # What the Supplier cell said, kept verbatim. This is exactly
                # what moving the cell into 'Dependencies/Notes' on the sheet
                # would have produced.
                misplaced,
                f"Paid date (as written): {raw_date}" if raw_date and not paid_date else None,
                f"Hours: {num(r[C_HOURS])}" if num(r[C_HOURS]) else None,
                f"Rate: {num(r[C_RATE])}" if num(r[C_RATE]) else None,
            ),
            "supplier": supplier,
            "invoice": text(r[C_INVOICE]),
            "paid_date": paid_date,
            "method": method_of(r[C_METHOD]),
            "qty": num(r[C_QTY]),
            "unit_cost": num(r[C_UNIT]),
            "vat": vat,
            "status": status_of(r[C_STATUS], sheet_total if is_paid else 0),
            # See the module docstring for why each amount is what it is.
            "quoted": sheet_total,
            "actual": ex_vat,
            "paid": sheet_total if is_paid else 0,
            "sheet_total": sheet_total,
            "source": "diary",
        })
    return out, warnings


def check_arithmetic(diary):
    """Verify actual x (1 + vat) reproduces the sheet's own Total column.

    If this ever fails, the row's Labour/Materials/VAT/Total cells disagree with
    each other in the spreadsheet and the import would silently invent a number.
    """
    bad = []
    for r in diary:
        expect = round(r["actual"] * (1 + r["vat"] / 100), 2)
        if abs(expect - r["sheet_total"]) > PENNY:
            bad.append(
                f"row {r['excel_row']} (week {r['week']}, {r['description']!r}): "
                f"ex-VAT {r['actual']} + {r['vat']}% = {expect} but the sheet's "
                f"Total column says {r['sheet_total']}"
            )
    return bad


def check_suppliers(diary):
    """Prove SUPPLIER_NAMES still describes the spreadsheet exactly.

    The whole mapping rests on one claim: the distinct descriptions of the
    Materials rows ARE the merchant list the owner confirmed. If someone adds
    a Materials row for a new merchant, or renames one, that claim stops being
    true and the import would quietly leave the row with no supplier. This
    turns that into a hard stop instead.
    """
    on_sheet = {name_key(r["description"]) for r in diary if r["category"] == "Materials"}
    declared = {name_key(n) for n in SUPPLIER_NAMES}
    bad = []
    for k in sorted(on_sheet - declared):
        rows = [r for r in diary if name_key(r["description"]) == k]
        bad.append(
            f"Materials row(s) {[r['excel_row'] for r in rows]} describe "
            f"{rows[0]['description']!r}, which is not in SUPPLIER_NAMES — add it "
            f"there (after checking with the owner) or the row imports with no supplier"
        )
    for k in sorted(declared - on_sheet):
        bad.append(
            f"SUPPLIER_NAMES lists {SUPPLIER_BY_KEY[k]!r} but no Materials row on "
            f"the sheet has that description any more — remove it or fix the sheet"
        )
    return bad


def parse_week_summary():
    """The Summary sheet's 'By Week (Totals)' block -> {week: total incl VAT}.

    Used as the independent expected value in the generated migration's own
    self-check, so the SQL proves itself against the spreadsheet.
    """
    ws = openpyxl.load_workbook(FILE1, data_only=True)["Summary"]
    out = {}
    for r in ws.iter_rows(values_only=True):
        if isinstance(r[0], (int, float)) and r[3] is not None:
            out[int(r[0])] = round(float(r[3]), 2)
    return out


# ----------------------------------------------------------------- lookups
def parse_lookups():
    ws = openpyxl.load_workbook(FILE1, data_only=True)["Lookups"]
    out = []
    for r in list(ws.iter_rows(values_only=True))[1:]:
        if not r[0]:
            continue
        out.append((str(r[0]).strip(), num(r[1]), num(r[2])))
    return out


# ----------------------------------------------------------------- report
def app_figures(diary):
    """Recompute the Overview cards exactly as lib/summary.ts would.

    Kept in step with buildSummary(): quoted and paid are summed straight from
    the stored columns, the total is derived as actual x (1 + vat/100), and
    cancelled rows are excluded.
    """
    active = [r for r in diary if r["status"] != "Cancelled"]
    total_quoted = sum(r["quoted"] for r in active)
    forecast = sum(r["actual"] * (1 + r["vat"] / 100) for r in active)
    paid = sum(r["paid"] for r in active)
    return {
        "total_quoted": round(total_quoted, 2),
        "forecast_total": round(forecast, 2),
        "variance": round(forecast - total_quoted, 2),
        "paid_to_date": round(paid, 2),
        "remaining_to_pay": round(forecast - paid, 2),
        "weeks_tracked": len({r["week"] for r in active}),
    }


def supplier_report(diary):
    """Rows and incl-VAT spend per merchant — what /suppliers will show.

    Cancelled rows are excluded, matching ACTIVE in lib/summary.ts and
    ACTIVE_PURCHASE in lib/purchases.ts.
    """
    named = {}
    for r in diary:
        if r["status"] == "Cancelled" or not r["supplier"]:
            continue
        d = named.setdefault(r["supplier"], {"rows": 0, "gross": 0.0})
        d["rows"] += 1
        d["gross"] += r["actual"] * (1 + r["vat"] / 100)
    return named


def print_week_report(diary, sheet_weeks):
    """Week-by-week: what the app will show vs what the spreadsheet says."""
    by_week = {}
    for r in diary:
        if r["status"] == "Cancelled":
            continue
        b = by_week.setdefault(r["week"], {"labour": 0.0, "materials": 0.0, "total": 0.0, "n": 0})
        incl = r["actual"] * (1 + r["vat"] / 100)
        if r["category"] == "Materials":
            b["materials"] += incl
        else:
            b["labour"] += incl
        b["total"] += incl
        b["n"] += 1

    print("\n  week  rows       app total     sheet total        diff")
    print("  " + "-" * 56)
    worst = 0.0
    for w in sorted(by_week):
        app = by_week[w]["total"]
        sheet = sheet_weeks.get(w, 0.0)
        diff = app - sheet
        worst = max(worst, abs(diff))
        flag = "" if abs(diff) <= PENNY else "   <-- MISMATCH"
        print(f"  {w:>4}  {by_week[w]['n']:>4}  {app:>14,.2f}  {sheet:>14,.2f}  {diff:>10.3f}{flag}")
    print(f"  worst week difference: £{worst:.4f}")
    return by_week


# ----------------------------------------------------------------- emit
def main():
    diary, warnings = parse_diary()
    sheet_weeks = parse_week_summary()

    problems = check_arithmetic(diary)
    if problems:
        print("ABORT — the spreadsheet's own arithmetic does not add up:")
        for p in problems:
            print("  " + p)
        raise SystemExit(1)

    problems = check_suppliers(diary)
    if problems:
        print("ABORT — SUPPLIER_NAMES no longer matches the spreadsheet:")
        for p in problems:
            print("  " + p)
        raise SystemExit(1)

    lookups = parse_lookups()
    entries = diary  # File 1 only — see the module docstring for why

    weeks = sorted({e["week"] for e in diary})
    fig = app_figures(diary)
    named = supplier_report(diary)
    with_supplier = sum(1 for r in diary if r["supplier"])

    # Only weeks that actually carry rows are worth checking; the sheet lists
    # empty weeks up to 32.
    expected_weeks = [(w, sheet_weeks[w]) for w in weeks if w in sheet_weeks]

    sql = [
        "-- RenovaTrack — rebuild from the Week-by-Week Plan alone, weeks 1-23.",
        "-- GENERATED by scripts/build_import_sql.py — do not hand-edit.",
        "--",
        f"-- Owner  : {USER_ID} (admin@pk.com)",
        f"-- Diary  : {len(diary)} rows, weeks {min(weeks)}-{max(weeks)}",
        "--          (46_Glenferrie_Rd_Renovation_Spend_Tracker_Updated.xlsx,",
        "--           'Week-by-Week Plan')",
        f"-- Trades : {len(lookups)} lookups (same file, 'Lookups')",
        f"-- Supplrs: {len(named)} merchants on {with_supplier} rows. The sheet's",
        "--          'Supplier' column is NOT the source: it has 8 non-empty cells",
        "--          and 7 are notes typed into the wrong column ('steels in',",
        "--          '£300 PAID FROM OWED'). The merchant names were typed into",
        "--          'Task / Description' on the Materials rows, and that is where",
        "--          this import reads them from, against a list confirmed by the",
        "--          owner (SUPPLIER_NAMES in the script). The 7 notes are moved",
        "--          into notes, where they belong. Re-run 0008 afterwards to",
        "--          reseed public.suppliers and link every purchase to one.",
        "-- Ledger : 0 rows — Renovation_Cost_Tracker-1.xlsx is NO LONGER",
        "--          imported. It is a different job: it names no address, its",
        "--          dates (2025-11-02 to 2026-01-25) stop a month before this",
        "--          project's week 1, and it ends with carpets and staging while",
        "--          week 1 here is 'clearance / back to brick'. It also broke the",
        "--          Price Tracker — its Unit Cost column holds the amount of each",
        "--          instalment, so a 416.05 deposit followed by a 3,586.00 balance",
        "--          to Wunda UFH read as a 761.9% price rise. The workbook stays",
        "--          in the repo; it is simply not imported.",
        "--",
        "-- Supersedes 0005, 0006 and 0007. Do NOT run any of those again.",
        "--",
        "-- target_budget is set to 0 because the spreadsheet's own 'Target Budget",
        "-- (incl. VAT)' cell is blank. The budget card and the '% of budget' bar",
        "-- hide themselves at 0 rather than comparing spend against another job's",
        "-- total, which is what the old 98,932.12 figure was.",
        "--",
        "-- Money mapping — actual_amount is EX-VAT because the app derives",
        "--   total_incl_vat = actual_amount x (1 + vat_rate/100)",
        "-- on every read (lib/calculations.ts). Storing an incl-VAT figure here",
        "-- is what made the old import show VAT twice.",
        "--   actual_amount = Labour Cost + Materials Cost   (ex-VAT)",
        "--   quoted_amount = Total incl. VAT                (the sheet's forecast)",
        "--   paid_amount   = Total incl. VAT, when a Paid Date is filled",
        "--",
        "-- After this runs the Overview tab should read:",
        f"--   Total Quoted      £{fig['total_quoted']:,.2f}",
        f"--   Actual Total      £{fig['forecast_total']:,.2f}   (= the sheet's Forecast Total)",
        f"--   Variance vs Quote £{fig['variance']:,.2f}",
        f"--   Paid to Date      £{fig['paid_to_date']:,.2f}   (= the sheet's 'Paid to date')",
        f"--   Remaining to Pay  £{fig['remaining_to_pay']:,.2f}  (= the sheet's 'Committed')",
        f"--   Weeks Tracked     {fig['weeks_tracked']}",
        "--",
        "-- Run in the Supabase SQL editor. Re-runnable: it deletes this user's",
        "-- existing copy of the project first. It refuses to commit if any week",
        "-- total disagrees with the spreadsheet.",
        "",
        "begin;",
        "",
        "do $$",
        "declare",
        f"  v_user    uuid := '{USER_ID}';",
        "  v_project uuid;",
        "begin",
        "  if not exists (select 1 from auth.users where id = v_user) then",
        "    raise exception 'User % not found in auth.users.', v_user;",
        "  end if;",
        "",
        "  -- Idempotency: clear any prior import of this project. Cascades to",
        "  -- expense_entries, project_weeks, and (from 0008) purchases with their",
        "  -- lines, payments and receipt rows.",
        f"  delete from public.projects where user_id = v_user and name = {q(PROJECT_NAME)};",
        "",
        "  -- 0008's suppliers and items sit ABOVE the project, so the delete above",
        "  -- does not reach them: without this, every merchant seeded from the old",
        "  -- File 2 import would linger on /suppliers and /items with no purchase",
        "  -- behind it. Aliases cascade; purchases.supplier_id and",
        "  -- purchase_lines.item_id are 'on delete set null', so nothing else",
        "  -- breaks. Re-run 0008 afterwards to reseed them from what is left.",
        "  delete from public.items     where user_id = v_user;",
        "  delete from public.suppliers where user_id = v_user;",
        "",
        "  insert into public.projects (user_id, name, target_budget, status, notes)",
        f"  values (v_user, {q(PROJECT_NAME)}, {TARGET_BUDGET}, 'active',",
        f"          'Imported from the Week-by-Week Plan, weeks 1-{max(weeks)}.')",
        "  returning id into v_project;",
        "",
        f"  insert into public.expense_entries ({COLS}) values",
    ]
    sql.append(",\n".join(row_sql(r) for r in entries) + ";")
    sql += ["", "  insert into public.trade_lookups (user_id, name, default_rate, default_markup_pct) values"]
    sql.append(
        ",\n".join(f"  (v_user, {q(n)}, {rate}, {mk})" for n, rate, mk in lookups)
        + "\n  on conflict (user_id, name) do update"
        + "\n    set default_rate = excluded.default_rate,"
        + "\n        default_markup_pct = excluded.default_markup_pct;"
    )
    sql += [
        "",
        "  insert into public.project_weeks (user_id, project_id, week_number, completion_pct)",
        f"  select v_user, v_project, generate_series(1, {max(weeks)}), 0",
        "  on conflict (project_id, week_number) do nothing;",
        "",
        f"  raise notice 'Imported project % ({len(diary)} diary rows, no ledger).', v_project;",
        "end $$;",
        "",
        "-- ------------------------------------------------------------------",
        "-- Self-check: every week's diary total must equal the spreadsheet's",
        "-- 'Summary' -> 'By Week (Totals)' figure. Any disagreement aborts the",
        "-- transaction, so a bad import can never be committed by accident.",
        "-- ------------------------------------------------------------------",
        "do $$",
        "declare",
        f"  v_user  uuid := '{USER_ID}';",
        "  r       record;",
        "  v_bad   int := 0;",
        "begin",
        "  for r in",
        "    with expected(week_number, sheet_total) as (values",
    ]
    sql.append(
        ",\n".join(f"      ({w}, {t}::numeric)" for w, t in expected_weeks)
        + "\n    ),"
    )
    sql += [
        "    got as (",
        "      select e.week_number,",
        "             sum(e.actual_amount * (1 + e.vat_rate / 100)) as app_total",
        "      from public.expense_entries e",
        "      join public.projects p on p.id = e.project_id",
        "      where p.user_id = v_user",
        f"        and p.name   = {q(PROJECT_NAME)}",
        "        and e.source = 'diary'",
        "        and e.status <> 'Cancelled'",
        "      group by e.week_number",
        "    )",
        "    select expected.week_number, expected.sheet_total,",
        "           coalesce(got.app_total, 0) as app_total",
        "    from expected left join got using (week_number)",
        "    order by expected.week_number",
        "  loop",
        "    if abs(r.app_total - r.sheet_total) > 0.01 then",
        "      v_bad := v_bad + 1;",
        "      raise warning 'Week %: app % <> spreadsheet %',",
        "        r.week_number, round(r.app_total, 2), r.sheet_total;",
        "    end if;",
        "  end loop;",
        "",
        "  if v_bad > 0 then",
        "    raise exception '% week(s) do not match the spreadsheet — nothing committed.', v_bad;",
        "  end if;",
        f"  raise notice 'All {len(expected_weeks)} weeks match the spreadsheet.';",
        "end $$;",
        "",
        "commit;",
        "-- rollback;  -- use instead of commit if the numbers look wrong",
        "",
        "-- ==================================================================",
        "-- NEXT STEP: run 0008_transaction_core.sql again, immediately.",
        "-- It is re-runnable. Its backfill was wiped by the project delete",
        "-- above (purchases.project_id is on delete cascade), and its suppliers",
        "-- and items were cleared on purpose so the old File 2 merchants do not",
        "-- linger. Until you re-run it, /suppliers and /items will be empty.",
        "-- ==================================================================",
        "",
        "-- ------------------------------------------------------------------",
        "-- Report. Compare these against the spreadsheet's Summary tab.",
        "-- 'Ledger rows' must now be 0.",
        "-- ------------------------------------------------------------------",
        "select 'Total Quoted'      as card, sum(quoted_amount)                          as value",
        "from public.expense_entries e join public.projects p on p.id = e.project_id",
        f"where p.user_id = '{USER_ID}' and p.name = {q(PROJECT_NAME)}",
        "  and e.source = 'diary' and e.status <> 'Cancelled'",
        "union all",
        "select 'Actual Total (incl. VAT)', sum(actual_amount * (1 + vat_rate / 100))",
        "from public.expense_entries e join public.projects p on p.id = e.project_id",
        f"where p.user_id = '{USER_ID}' and p.name = {q(PROJECT_NAME)}",
        "  and e.source = 'diary' and e.status <> 'Cancelled'",
        "union all",
        "select 'Paid to Date', sum(paid_amount)",
        "from public.expense_entries e join public.projects p on p.id = e.project_id",
        f"where p.user_id = '{USER_ID}' and p.name = {q(PROJECT_NAME)}",
        "  and e.source = 'diary' and e.status <> 'Cancelled'",
        "union all",
        "select 'Weeks Tracked', count(distinct week_number)",
        "from public.expense_entries e join public.projects p on p.id = e.project_id",
        f"where p.user_id = '{USER_ID}' and p.name = {q(PROJECT_NAME)}",
        "  and e.source = 'diary' and e.status <> 'Cancelled'",
        "union all",
        "select 'Diary rows', count(*)",
        "from public.expense_entries e join public.projects p on p.id = e.project_id",
        f"where p.user_id = '{USER_ID}' and p.name = {q(PROJECT_NAME)} and e.source = 'diary'",
        "union all",
        "select 'Ledger rows', count(*)",
        "from public.expense_entries e join public.projects p on p.id = e.project_id",
        f"where p.user_id = '{USER_ID}' and p.name = {q(PROJECT_NAME)} and e.source = 'ledger';",
        "",
    ]

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text("\n".join(sql), encoding="utf-8")

    # ------------------------------------------------------------- summary
    print(f"wrote {OUT.relative_to(ROOT)}")
    print(f"  source : {FILE1.name}")
    print(f"  diary  : {len(diary)} rows (weeks {min(weeks)}-{max(weeks)})")
    print("  ledger : 0 rows — Renovation_Cost_Tracker-1.xlsx is a different job")
    print(f"  lookups: {len(lookups)}")
    print(f"  paid   : {sum(1 for r in diary if r['paid'] > 0)} rows")
    print("  every row's ex-VAT + VAT reproduces the sheet's Total column")

    print(f"\nSuppliers: {len(named)} distinct on {with_supplier} of "
          f"{len(diary)} rows ({len(diary) - with_supplier} rows have no merchant "
          "behind them — labour and subcontractors)")
    for name in sorted(named, key=lambda n: (-named[n]["rows"], n.lower())):
        d = named[name]
        print(f"  {name:<28} {d['rows']:>2} row(s)  £{d['gross']:>10,.2f} incl VAT")

    print("\nOverview cards this import will produce:")
    print(f"  Target Budget      £{TARGET_BUDGET:,.2f}   (the sheet's own cell is blank; "
          "the card hides itself at 0)")
    print(f"  Total Quoted       £{fig['total_quoted']:,.2f}")
    print(f"  Actual Total       £{fig['forecast_total']:,.2f}")
    print(f"  Variance vs Quote  £{fig['variance']:,.2f}")
    print(f"  Paid to Date       £{fig['paid_to_date']:,.2f}")
    print(f"  Remaining to Pay   £{fig['remaining_to_pay']:,.2f}")
    print(f"  Weeks Tracked      {fig['weeks_tracked']}")

    print_week_report(diary, sheet_weeks)

    for w in warnings:
        print(f"  note: {w}")


if __name__ == "__main__":
    main()
